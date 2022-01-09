from calendar import monthrange
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles.alignment import Alignment
import psycopg2, random, locale
locale.setlocale(locale.LC_ALL, '')
year = 2021
month = 12

class Shift:

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.lst = {}



    def parseMonth(self,year, month):
        self.days = monthrange(year, month)[1]
        weekday = []
        weekend = []
        for day in range(1, self.days + 1):
            which = datetime(year, month, day).isoweekday()
            if which in [6, 7]:
                weekend.append(day)
            else:
                weekday.append(day)
        self.weekday = weekday
        self.weekend = weekend
        self._change = {}
        for i in range(1,8):
            self._change[i] = datetime(year,month,i).strftime("%A")
            
            
    def createExcel(self):
        feautres = ["SIRA", "TARİH", "GÜN", "GÖREVİ", "ADI SOYADI", "ÜNVANI", "NÖBET\nSAATLERİ"]

        for index, column in enumerate(feautres):
            self.ws[f"{chr(index + 65)}1"] = column

        for index, value in enumerate(range(2, (self.days * 2) + 2, 2)):
            self.ws.merge_cells(f"A{value}:A{value + 1}")
            self.ws[f"A{value}"] = index + 1
            self.ws.merge_cells(f"B{value}:B{value + 1}")
            self.ws[f"B{value}"] = datetime(year, month, index + 1).date()
            self.ws.merge_cells(f"C{value}:C{value + 1}")
            self.ws[f"C{value}"] = datetime(year, month, index + 1).strftime("%A")
            self.ws[f"D{value}"] = "Work 1"
            self.ws[f"D{value + 1}"] = "Work 2"
            self.ws[f"G{value}"] = "17:00-08:00"
            self.ws[f"G{value + 1}"] = "08:00-08:00"
            self.ws[f"F{value}"] = "Title 1"
            self.ws[f"F{value + 1}"] = "Title 2"

        self.wb.save("sample.xlsx")

    def pullData(self, tabloName):
        conn = psycopg2.connect(host="localhost", database="excelSheet", user="postgres", password="root")
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM {tabloName} ORDER BY weekend ASC LIMIT {len(self.weekend)}")
        weekend = cursor.fetchall()
        cursor.execute(f"SELECT * FROM {tabloName} ORDER BY weekday ASC LIMIT {len(self.weekday)}")
        weekday = cursor.fetchall()
        for name,end,day in weekend:
            cursor.execute(f"UPDATE {tabloName} SET weekend = {end+1} where name = '{name}'")
        for name,end,day in weekday:
            cursor.execute(f"UPDATE {tabloName} SET weekday = {day+1} where name = '{name}'")
        conn.commit()


        cursor.close()
        conn.close()
        return (weekday, weekend)

    def personelPut(self,tabloName,offset=0):
        print(self._change)
        dayPersonel, endPersonel = shift.pullData(tabloName)
        for day in self.weekend:
            index = -1
            while True:
                person = endPersonel[index][0]
                if person not in self.lst:
                    self.ws[f"E{day * 2 + offset}"] = endPersonel.pop(index)[0]
                    break
                if self.lst[person] != self._change[((day%7)+1)]:
                    self.ws[f"E{day*2+offset}"] = endPersonel.pop(index)[0]
                    break
                else:
                    index-=1
        for day in self.weekday:
            index = -1
            while True:
                person = dayPersonel[index][0]
                if person not in self.lst:
                    self.ws[f"E{day * 2 + offset}"] = dayPersonel.pop(index)[0]
                    break
                if self.lst[person] != self._change[((day%7)+1)]:
                    self.ws[f"E{day*2+offset}"] = dayPersonel.pop(index)[0]
                    break
                else:
                    index-=1

        self.wb.save("sample.xlsx")

    def offDay(self,name,day):
        self.lst[name] = day

    def offset(self):
        for i in range(8):
            for k in range(1,61):
                self.ws[f"{chr(i+65)}{k}"].alignment = Alignment(horizontal="center",vertical="center")
        self.wb.save("sample.xlsx")

    def addPerson(self, name):
        conn = psycopg2.connect(host="localhost", database="excelSheet", user="postgres", password="root")
        cursor = conn.cursor()
        cursor.execute(f"INSERT INTO WORK1 VALUES('{name}',(SELECT AVG(WEEKEND) FROM WORK1), (SELECT AVG(WEEKDAY) FROM WORK1))")
        cursor.execute(f"INSERT INTO WORK2 VALUES('{name}',(SELECT AVG(WEEKEND) FROM WORK2), (SELECT AVG(WEEKDAY) FROM WORK2))")
        conn.commit()

    def removePerson(self, name):
        conn = psycopg2.connect(host="localhost", database="excelSheet", user="postgres", password="root")
        cursor = conn.cursor()
        cursor.execute(f"DELETE FROM WORK1 WHERE name = '{name}'")
        cursor.execute(f"DELETE FROM WORK2 WHERE name = '{name}'")
        conn.commit()
        
        
shift = Shift()
shift.parseMonth(year, month)
shift.createExcel()
# shift.offDay()
shift.personelPut('WORK1')
shift.personelPut('WORK2',offset=1)
shift.offset()
